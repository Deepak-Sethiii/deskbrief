"""Prints the exact hand-build steps for templates/deskbrief_template.xlsm.

Generated from the layout constants in excel.py so the instructions and the
writer cannot drift apart. Run: python -m src.cli template-help
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from src.report.excel import (
    CHART_ANCHOR,
    CHART_TITLE_CELL,
    COMMENTARY_TITLE_CELL,
    HEADLINE_HEADERS,
    HEADLINES_FIRST_DATA_ROW,
    REQUIRED_SHEETS,
    SUMMARY_FIRST_DATA_ROW,
    SUMMARY_HEADERS,
    SUMMARY_STAMP_CELL,
    TEMPLATE_PATH,
)

RULE = "=" * 78


def render() -> str:
    lines: list[str] = []
    add = lines.append

    add(RULE)
    add("ONE-TIME SETUP: build templates/deskbrief_template.xlsm by hand")
    add(RULE)
    add("")
    add("WHY BY HAND: openpyxl cannot create VBA. It can only PRESERVE a")
    add("vbaProject that is already inside a file it opened with keep_vba=True.")
    add("So the macro-enabled template has to be created once in Excel itself.")
    add("After this, you never touch it again -- every run copies from it.")
    add("")
    add(f"Target file: {TEMPLATE_PATH}")
    add("")

    add("-" * 78)
    add("STEP 1 - New workbook, four sheets")
    add("-" * 78)
    add("  1.1  Open Excel -> Blank workbook.")
    add("  1.2  Rename Sheet1 to exactly:  Summary")
    add("       (right-click the tab -> Rename. Names are case-sensitive here.)")
    add("  1.3  Add three more sheets with the + button and rename them exactly:")
    for name in REQUIRED_SHEETS[1:]:
        add(f"         {name}")
    add(f"  1.4  Tab order should read: {' | '.join(REQUIRED_SHEETS)}")
    add("       Delete any other sheet. The writer fails loudly if one is missing.")
    add("")

    add("-" * 78)
    add("STEP 2 - Summary sheet: header row and formatting")
    add("-" * 78)
    add(f"  2.1  Cell {SUMMARY_STAMP_CELL} : leave empty. The pipeline writes the")
    add("       'DeskBrief - generated <timestamp>' stamp here. Set it Bold, size 14.")
    header_row = SUMMARY_FIRST_DATA_ROW - 1
    add(f"  2.2  Type these headers across row {header_row}, one per column:")
    add("")
    add("         Cell   Header               Number format for the column below")
    add("         -----  -------------------  ----------------------------------")
    for index, (header, fmt) in enumerate(SUMMARY_HEADERS, start=1):
        letter = get_column_letter(index)
        add(f"         {letter}{header_row:<5} {header:<20} {fmt}")
    add("")
    add(f"  2.3  Select row {header_row} -> Bold, white text on a dark fill, centre it,")
    add("       and add a bottom border. Freeze panes below it:")
    add(f"       click A{SUMMARY_FIRST_DATA_ROW} -> View -> Freeze Panes -> Freeze Panes.")
    add(f"  2.4  Apply the number format from the table above to each column from")
    add(f"       row {SUMMARY_FIRST_DATA_ROW} down (select the whole column, Ctrl+1).")
    add("       THIS MATTERS: the pipeline writes raw fractions (0.0254), not")
    add("       '2.54%' strings. Without the 0.00% format they will read as 0.03.")
    add("  2.5  Widen the columns so nothing shows as ####.")
    add(f"  2.6  Optional: select A{header_row}:"
        f"{get_column_letter(len(SUMMARY_HEADERS))}{header_row} -> Data -> Filter.")
    add("")

    add("-" * 78)
    add("STEP 3 - Headlines sheet")
    add("-" * 78)
    hl_header_row = HEADLINES_FIRST_DATA_ROW - 1
    add(f"  3.1  Type these headers across row {hl_header_row}:")
    for index, (header, fmt) in enumerate(HEADLINE_HEADERS, start=1):
        letter = get_column_letter(index)
        add(f"         {letter}{hl_header_row}   {header:<12} format: {fmt}")
    add("  3.2  Bold that row. Widen column D (Title) to ~80 and E (URL) to ~50.")
    add("  3.3  Format column A as yyyy-mm-dd hh:mm.")
    add("")

    add("-" * 78)
    add("STEP 4 - Charts and Commentary sheets")
    add("-" * 78)
    add(f"  4.1  Charts: leave {CHART_TITLE_CELL} empty (the pipeline writes the")
    add(f"       title there). The chart PNG is anchored at {CHART_ANCHOR}.")
    add("       Bold A1, size 12. Nothing else is needed on this sheet.")
    add(f"  4.2  Commentary: leave {COMMENTARY_TITLE_CELL} empty (title written")
    add("       there). Bold it, size 14. Widen column B to ~100 and set it to")
    add("       Wrap Text so the bullets are readable.")
    add("")

    add("-" * 78)
    add("STEP 5 - Save as macro-enabled")
    add("-" * 78)
    add("  5.1  File -> Save As -> browse to the repo's templates\\ folder.")
    add("  5.2  Set 'Save as type' to:  Excel Macro-Enabled Workbook (*.xlsm)")
    add("       This is the step people miss. A .xlsx cannot hold VBA at all.")
    add(f"  5.3  Filename: {TEMPLATE_PATH.name}")
    add("  5.4  Save. Keep it open for the next step.")
    add("")

    add("-" * 78)
    add("STEP 6 - Import vba/Refresh.bas")
    add("-" * 78)
    add("  6.1  If you do not see the Developer tab: File -> Options ->")
    add("       Customize Ribbon -> tick 'Developer' on the right -> OK.")
    add("  6.2  Press Alt+F11 to open the VBA editor.")
    add("  6.3  In the Project pane on the left, click VBAProject ("
        f"{TEMPLATE_PATH.name}) so it is selected.")
    add("  6.4  File -> Import File...  (Ctrl+M)")
    add("  6.5  Choose the repo's  vba\\Refresh.bas  and click Open.")
    add("  6.6  A 'Refresh' module appears under Modules. Double-click it and")
    add("       confirm you can see Public Sub RefreshDeskBrief().")
    add("  6.7  Alt+Q to return to Excel.")
    add("")

    add("-" * 78)
    add("STEP 7 - Add the button")
    add("-" * 78)
    add("  7.1  Go to the Summary sheet.")
    add("  7.2  Insert -> Shapes -> Rounded Rectangle. Draw it somewhere clear of")
    add("       the data, e.g. over columns O-P near row 1.")
    add("  7.3  Type the label:  Refresh DeskBrief")
    add("  7.4  Right-click the shape -> Assign Macro...")
    add("  7.5  Select  RefreshDeskBrief  -> OK.")
    add("       (If the list is empty, the import in step 6 did not take.)")
    add("  7.6  Click off the shape. One left-click now runs the pipeline.")
    add("")

    add("-" * 78)
    add("STEP 8 - Save and verify")
    add("-" * 78)
    add("  8.1  Ctrl+S. If Excel offers to save as .xlsx, say NO and keep .xlsm.")
    add("  8.2  Close Excel completely.")
    add("  8.3  Verify from the repo root:")
    add("         python -m src.cli verify-template")
    add("       It checks the four sheet names and that vbaProject.bin survived.")
    add("")
    add(RULE)
    add("NOTE ON TRUST: the first time you open the generated output workbook,")
    add("Excel shows a yellow 'SECURITY WARNING - Macros have been disabled'")
    add("bar. Click 'Enable Content'. Alternatively add the repo's output\\")
    add("folder as a Trusted Location (File -> Options -> Trust Center ->")
    add("Trust Center Settings -> Trusted Locations).")
    add(RULE)

    return "\n".join(lines)


if __name__ == "__main__":  # pragma: no cover
    print(render())
