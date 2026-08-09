"""Write the analyst workbook.

THE ONE RULE IN THIS MODULE: the template owns formatting, this code owns
values. openpyxl cannot create VBA -- it can only *preserve* a vbaProject that
already exists in a file opened with keep_vba=True. So we never create a sheet,
never set a font, fill, width, border or number format, and never touch a
header row. We open templates/deskbrief_template.xlsm, drop in cell values and
one image, and save under a new name. If a number looks wrong in Excel, fix the
template's number format, not this file.

Every run starts from the pristine template rather than from the previous
output, so stale rows from a longer prior run cannot survive into a shorter one.

Excel holds an exclusive write lock on an open workbook, so writing in place is
impossible by construction. We write output/deskbrief_YYYYMMDD_HHMM.xlsm and
record its absolute path in output/latest.txt; vba/Refresh.bas reads that file
and swaps the user over. Do not attempt to work around the lock.
"""

from __future__ import annotations

import datetime as dt
import logging
import math
from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from src.paths import LATEST_POINTER_NAME, OUTPUT_ROOT, TEMPLATE_PATH

log = logging.getLogger(__name__)

SHEET_SUMMARY = "Summary"
SHEET_CHARTS = "Charts"
SHEET_HEADLINES = "Headlines"
SHEET_COMMENTARY = "Commentary"
REQUIRED_SHEETS = (SHEET_SUMMARY, SHEET_CHARTS, SHEET_HEADLINES, SHEET_COMMENTARY)

# --- layout contract -------------------------------------------------------
# Single source of truth shared by the writer and the template build
# instructions (docs/TEMPLATE_SETUP.md is generated from these constants).
SUMMARY_HEADERS: tuple[tuple[str, str], ...] = (
    # (header text, number format the TEMPLATE should carry for that column)
    ("Ticker", "General"),
    ("Sector", "General"),
    ("Ccy", "General"),
    ("Last Close", "#,##0.00"),
    ("1D", "0.00%"),
    ("5D", "0.00%"),
    ("21D", "0.00%"),
    ("Vol 30D (ann.)", "0.0%"),
    ("52W High", "#,##0.00"),
    ("vs 52W High", "0.0%"),
    ("Mkt Cap (bn)", "#,##0"),
    ("P/E", "#,##0.0"),
    ("EV/EBITDA", "#,##0.0"),
)
SUMMARY_FIRST_DATA_ROW = 3  # row 1 = title/stamp, row 2 = headers
SUMMARY_STAMP_CELL = "A1"

HEADLINE_HEADERS: tuple[tuple[str, str], ...] = (
    ("Published", "yyyy-mm-dd hh:mm"),
    ("Source", "General"),
    ("Ticker", "General"),
    ("Title", "General"),
    ("URL", "General"),
)
HEADLINES_FIRST_DATA_ROW = 2
MAX_HEADLINES = 60

CHART_TITLE_CELL = "A1"
CHART_ANCHOR = "A3"

COMMENTARY_TITLE_CELL = "A1"
COMMENTARY_FIRST_ROW = 3

# Comps columns pulled onto the Summary sheet, in header order.
SUMMARY_COLUMNS = (
    "sector", "currency", "last_close",
    "ret_1d", "ret_5d", "ret_21d",
    "realised_vol_30d", "high_52w", "pct_from_52w_high",
    "market_cap_bn", "pe", "ev_ebitda",
)


class TemplateMissingError(FileNotFoundError):
    """Raised when the hand-built .xlsm template is absent or malformed."""


def _cell_value(value: Any) -> Any:
    """Coerce to something openpyxl can store. NaN/NaT/inf become blank cells."""
    if value is None:
        return None
    if isinstance(value, float):
        return None if (math.isnan(value) or math.isinf(value)) else value
    if isinstance(value, (dt.datetime, dt.date, str, int, bool)):
        return value
    if pd.isna(value):
        return None
    if hasattr(value, "item"):  # numpy scalar
        try:
            return value.item()
        except Exception:  # noqa: BLE001
            return str(value)
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return str(value)


def _check_template(path: Path) -> None:
    if not path.exists():
        raise TemplateMissingError(
            f"{path} not found.\n"
            "openpyxl cannot create a macro-enabled workbook -- it can only preserve\n"
            "VBA that already exists. Build the template once by hand:\n"
            "    python -m src.cli template-help\n"
            "or run the bootstrapper:  python tools/build_template.py"
        )


def write_workbook(
    comps: pd.DataFrame,
    headlines: pd.DataFrame | None = None,
    commentary: Mapping[str, Any] | None = None,
    chart_path: str | Path | None = None,
    *,
    template: str | Path | None = None,
    output_dir: str | Path | None = None,
    generated_at: dt.datetime | None = None,
) -> Path:
    """Fill the template and save a timestamped .xlsm. Returns the output path."""
    template_path = Path(template) if template else TEMPLATE_PATH
    out_dir = Path(output_dir) if output_dir else OUTPUT_ROOT
    stamp = generated_at or dt.datetime.now()

    _check_template(template_path)
    out_dir.mkdir(parents=True, exist_ok=True)

    # keep_vba=True is the whole reason the template exists: it carries the
    # vbaProject.bin through untouched. Without it the macro is silently lost.
    workbook = load_workbook(template_path, keep_vba=True)

    missing = [s for s in REQUIRED_SHEETS if s not in workbook.sheetnames]
    if missing:
        raise TemplateMissingError(
            f"{template_path} is missing sheet(s) {missing}. "
            f"Expected exactly: {list(REQUIRED_SHEETS)}"
        )

    _write_summary(workbook[SHEET_SUMMARY], comps, stamp)
    _write_headlines(workbook[SHEET_HEADLINES], headlines)
    _write_commentary(workbook[SHEET_COMMENTARY], commentary, stamp)
    _write_chart(workbook[SHEET_CHARTS], chart_path, stamp)

    out_path = out_dir / f"deskbrief_{stamp:%Y%m%d_%H%M}.xlsm"
    workbook.save(out_path)
    workbook.close()

    absolute = out_path.resolve()
    # The pointer lives beside the workbook it points at, so a run targeting a
    # different output_dir (tests) cannot clobber the real one the macro reads.
    pointer = out_dir / LATEST_POINTER_NAME
    # Absolute path, because the macro resolves it from whatever directory Excel is in.
    pointer.write_text(str(absolute) + "\n", encoding="utf-8")

    log.info("workbook written: %s", absolute)
    log.info("latest pointer  : %s", pointer.resolve())
    return absolute


def _write_summary(sheet, comps: pd.DataFrame, stamp: dt.datetime) -> None:
    sheet[SUMMARY_STAMP_CELL] = f"DeskBrief - generated {stamp:%Y-%m-%d %H:%M}"

    row = SUMMARY_FIRST_DATA_ROW
    for ticker, record in comps.iterrows():
        sheet.cell(row=row, column=1, value=_cell_value(ticker))
        for offset, column in enumerate(SUMMARY_COLUMNS, start=2):
            sheet.cell(row=row, column=offset, value=_cell_value(record.get(column)))
        row += 1

    log.info("Summary: %d names written from row %d", len(comps), SUMMARY_FIRST_DATA_ROW)


def _write_headlines(sheet, headlines: pd.DataFrame | None) -> None:
    if headlines is None or headlines.empty:
        log.warning("Headlines: nothing to write")
        return

    frame = headlines.head(MAX_HEADLINES)
    row = HEADLINES_FIRST_DATA_ROW
    for _, record in frame.iterrows():
        values = (
            record.get("published_at"),
            record.get("source"),
            record.get("ticker"),
            record.get("title"),
            record.get("url"),
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=_cell_value(value))
        row += 1

    log.info("Headlines: %d rows written", len(frame))


def _write_commentary(sheet, commentary: Mapping[str, Any] | None, stamp: dt.datetime) -> None:
    sheet[COMMENTARY_TITLE_CELL] = f"Market commentary - {stamp:%Y-%m-%d %H:%M}"

    if not commentary:
        sheet.cell(row=COMMENTARY_FIRST_ROW, column=1, value="No commentary generated.")
        log.warning("Commentary: nothing to write")
        return

    row = COMMENTARY_FIRST_ROW
    sheet.cell(row=row, column=1, value="Market tone")
    sheet.cell(row=row, column=2, value=_cell_value(commentary.get("market_tone")))
    row += 2

    for label, key in (("Key points", "bullets"), ("Watch items", "watch_items")):
        sheet.cell(row=row, column=1, value=label)
        row += 1
        items: Sequence[str] = commentary.get(key) or []
        for item in items:
            sheet.cell(row=row, column=2, value=_cell_value(item))
            row += 1
        row += 1

    source = commentary.get("source")
    if source:
        sheet.cell(row=row, column=1, value="Source")
        sheet.cell(row=row, column=2, value=_cell_value(source))

    log.info("Commentary: tone + %d bullets + %d watch items",
             len(commentary.get("bullets") or []), len(commentary.get("watch_items") or []))


def _write_chart(sheet, chart_path: str | Path | None, stamp: dt.datetime) -> None:
    sheet[CHART_TITLE_CELL] = f"Normalised price performance - {stamp:%Y-%m-%d}"

    if not chart_path:
        log.warning("Charts: no image supplied")
        return
    path = Path(chart_path)
    if not path.exists():
        log.warning("Charts: image %s does not exist", path)
        return

    # Images are the one non-value thing we add; the template still owns the
    # sheet, we only anchor a picture onto it.
    sheet.add_image(XLImage(str(path)), CHART_ANCHOR)
    log.info("Charts: embedded %s at %s", path.name, CHART_ANCHOR)


def verify_template(path: str | Path | None = None) -> dict[str, Any]:
    """Check the hand-built template before a run depends on it.

    Confirms the four sheets exist and that vbaProject.bin actually survived the
    save -- the usual failure is saving as .xlsx, which drops the macro silently.
    """
    import zipfile

    template_path = Path(path) if path else TEMPLATE_PATH
    _check_template(template_path)

    with zipfile.ZipFile(template_path) as archive:
        names = archive.namelist()
    has_vba = "xl/vbaProject.bin" in names

    workbook = load_workbook(template_path, keep_vba=True)
    sheets = list(workbook.sheetnames)
    workbook.close()

    missing = [s for s in REQUIRED_SHEETS if s not in sheets]
    return {
        "path": str(template_path.resolve()),
        "sheets": sheets,
        "missing_sheets": missing,
        "has_vba": has_vba,
        # Missing sheets are fatal -- the writer cannot place values. Missing VBA
        # is not: the workbook still builds, you just lose the Refresh button.
        "ok": not missing,
        "fully_configured": not missing and has_vba,
    }


def summary_column_letters() -> list[str]:
    """Column letters for the Summary headers -- used by the setup instructions."""
    return [get_column_letter(i) for i in range(1, len(SUMMARY_HEADERS) + 1)]
