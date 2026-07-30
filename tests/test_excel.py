"""Stage 4 coverage.

The load-bearing test here is test_vba_survives_a_write: openpyxl cannot create
VBA, only preserve it, so the whole design rests on keep_vba=True carrying
vbaProject.bin from template to output. We prove that by injecting a marker
vbaProject into a copy of the template and asserting it comes out the far side
byte-identical.
"""

from __future__ import annotations

import datetime as dt
import shutil
import zipfile

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.report.excel import (
    SUMMARY_FIRST_DATA_ROW,
    TEMPLATE_PATH,
    TemplateMissingError,
    verify_template,
    write_workbook,
)

pytestmark = pytest.mark.skipif(
    not TEMPLATE_PATH.exists(),
    reason="templates/deskbrief_template.xlsm not built yet (see: cli template-help)",
)

VBA_MARKER = b"DESKBRIEF-FAKE-VBA-PROJECT-\x00\x01\x02"


@pytest.fixture()
def comps():
    return pd.DataFrame(
        {
            "sector": ["Energy", None],
            "currency": ["INR", None],
            "last_close": [1275.9, 748.2],
            "ret_1d": [0.0065, 0.0174],
            "ret_5d": [-0.0099, -0.0066],
            "ret_21d": [-0.0139, float("nan")],
            "realised_vol_30d": [0.175, 0.2537],
            "high_52w": [1584.97, 996.42],
            "pct_from_52w_high": [-0.195, -0.2491],
            "market_cap_bn": [17496.13, float("nan")],
            "pe": [23.4, float("nan")],
            "ev_ebitda": [11.47, float("nan")],
        },
        index=pd.Index(["RELIANCE.NS", "HDFCBANK.NS"], name="ticker"),
    )


@pytest.fixture()
def headlines():
    return pd.DataFrame(
        [
            {
                "published_at": dt.datetime(2026, 7, 30, 9, 15),
                "source": "Example Wire",
                "ticker": "RELIANCE.NS",
                "title": "Reliance posts higher Q1 profit",
                "url": "https://example.com/a",
            },
            {
                "published_at": None,
                "source": "Example Wire",
                "ticker": None,
                "title": "Undated market story",
                "url": "https://example.com/b",
            },
        ]
    )


def test_summary_values_land_in_the_right_cells(comps, tmp_path):
    out = write_workbook(comps, output_dir=tmp_path, generated_at=dt.datetime(2026, 7, 31, 1, 30))
    sheet = load_workbook(out, keep_vba=True)["Summary"]

    assert sheet["A1"].value == "DeskBrief - generated 2026-07-31 01:30"
    row = SUMMARY_FIRST_DATA_ROW
    assert sheet.cell(row=row, column=1).value == "RELIANCE.NS"
    assert sheet.cell(row=row, column=2).value == "Energy"
    assert sheet.cell(row=row, column=4).value == pytest.approx(1275.9)
    # raw fraction, NOT a preformatted "0.65%" string -- the template formats it
    assert sheet.cell(row=row, column=5).value == pytest.approx(0.0065)


def test_nan_becomes_a_blank_cell_not_the_text_nan(comps, tmp_path):
    out = write_workbook(comps, output_dir=tmp_path)
    sheet = load_workbook(out, keep_vba=True)["Summary"]
    hdfc_row = SUMMARY_FIRST_DATA_ROW + 1
    assert sheet.cell(row=hdfc_row, column=1).value == "HDFCBANK.NS"
    for column in (7, 11, 12, 13):  # ret_21d, mkt cap, pe, ev/ebitda
        assert sheet.cell(row=hdfc_row, column=column).value is None


def test_headlines_are_written(comps, headlines, tmp_path):
    out = write_workbook(comps, headlines, output_dir=tmp_path)
    sheet = load_workbook(out, keep_vba=True)["Headlines"]
    assert sheet["B2"].value == "Example Wire"
    assert sheet["C2"].value == "RELIANCE.NS"
    assert sheet["D2"].value == "Reliance posts higher Q1 profit"
    assert sheet["C3"].value is None  # untagged headline keeps a blank ticker


def test_commentary_is_written(comps, tmp_path):
    commentary = {
        "market_tone": "Constructive",
        "bullets": ["Point one", "Point two"],
        "watch_items": ["Watch one"],
        "source": "unit-test",
    }
    out = write_workbook(comps, commentary=commentary, output_dir=tmp_path)
    sheet = load_workbook(out, keep_vba=True)["Commentary"]
    flat = [c.value for row in sheet.iter_rows() for c in row if c.value]
    assert "Constructive" in flat
    assert "Point one" in flat and "Point two" in flat
    assert "Watch one" in flat


def test_output_filename_and_latest_pointer(comps, tmp_path):
    stamp = dt.datetime(2026, 7, 31, 14, 5)
    out = write_workbook(comps, output_dir=tmp_path, generated_at=stamp)
    assert out.name == "deskbrief_20260731_1405.xlsm"
    assert out.is_absolute()

    # the pointer is written beside the workbook, not into the real output/
    pointer = tmp_path / "latest.txt"
    assert pointer.read_text(encoding="utf-8").strip() == str(out)


def test_every_run_starts_from_the_template_so_stale_rows_cannot_survive(comps, tmp_path):
    """A long run followed by a short one must not leave orphaned rows behind."""
    write_workbook(comps, output_dir=tmp_path)
    shorter = comps.iloc[:1]
    out = write_workbook(shorter, output_dir=tmp_path,
                         generated_at=dt.datetime(2026, 7, 31, 23, 59))
    sheet = load_workbook(out, keep_vba=True)["Summary"]
    assert sheet.cell(row=SUMMARY_FIRST_DATA_ROW, column=1).value == "RELIANCE.NS"
    assert sheet.cell(row=SUMMARY_FIRST_DATA_ROW + 1, column=1).value is None


def test_missing_template_raises_with_actionable_advice(comps, tmp_path):
    with pytest.raises(TemplateMissingError, match="template-help"):
        write_workbook(comps, template=tmp_path / "nope.xlsm", output_dir=tmp_path)


def test_template_missing_a_sheet_is_rejected(comps, tmp_path):
    broken = tmp_path / "broken.xlsm"
    shutil.copy(TEMPLATE_PATH, broken)
    workbook = load_workbook(broken, keep_vba=True)
    del workbook["Commentary"]
    workbook.save(broken)

    with pytest.raises(TemplateMissingError, match="Commentary"):
        write_workbook(comps, template=broken, output_dir=tmp_path)


def _inject_fake_vba(source, destination) -> None:
    """Copy an xlsm, adding a marker vbaProject.bin and its content type."""
    with zipfile.ZipFile(source) as src:
        items = [(i, src.read(i.filename)) for i in src.infolist()]

    with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as dst:
        for info, payload in items:
            if info.filename == "[Content_Types].xml":
                text = payload.decode("utf-8")
                if "vbaProject" not in text:
                    override = (
                        '<Override PartName="/xl/vbaProject.bin" '
                        'ContentType="application/vnd.ms-office.vbaProject"/>'
                    )
                    text = text.replace("</Types>", override + "</Types>")
                payload = text.encode("utf-8")
            dst.writestr(info, payload)
        dst.writestr("xl/vbaProject.bin", VBA_MARKER)


def test_vba_survives_a_write(comps, tmp_path):
    """The core Stage 4 claim: keep_vba=True carries the macro into the output."""
    vba_template = tmp_path / "with_vba.xlsm"
    _inject_fake_vba(TEMPLATE_PATH, vba_template)

    with zipfile.ZipFile(vba_template) as archive:
        assert archive.read("xl/vbaProject.bin") == VBA_MARKER

    out = write_workbook(comps, template=vba_template, output_dir=tmp_path)

    with zipfile.ZipFile(out) as archive:
        assert "xl/vbaProject.bin" in archive.namelist(), "keep_vba dropped the macro"
        assert archive.read("xl/vbaProject.bin") == VBA_MARKER, "macro was mangled"


def test_verify_template_reports_vba_state(tmp_path):
    vba_template = tmp_path / "with_vba.xlsm"
    _inject_fake_vba(TEMPLATE_PATH, vba_template)

    without = verify_template(TEMPLATE_PATH)
    assert without["ok"] and not without["missing_sheets"]

    with_vba = verify_template(vba_template)
    assert with_vba["has_vba"] and with_vba["fully_configured"]
